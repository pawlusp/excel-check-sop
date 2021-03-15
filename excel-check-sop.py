from openpyxl import load_workbook, Workbook
import pandas as pd

wb1 = load_workbook("SOPs_search.xlsx")  # Work Book
ws1 = wb1.get_sheet_by_name('Search')  # Work Sheet for search
column = ws1['A']  # Column
column_list = []
for x in range(len(column)):    #List of items to search    
    value1 = column[x].value
    if value1 == None:
        pass
    else:
        column_list.append(value1)
#print(column_list)

ws2 = wb1.get_sheet_by_name('Data') # Work Sheet for data

columns_rows =[]
rows = []

def search_value_in_column(ws2, search_string, column="A"): #Search for items in data
    for row in range(1, ws2.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws2[coordinate].value == search_string:
            return column, row
    return column, None


for v in column_list:
    cr = search_value_in_column(ws2,v)
    cr = list(cr)
    if None in cr:
        pass
    else:
        columns_rows.append(cr)

#print(len(columns_rows))

#print(columns_rows)
for row in range(len(columns_rows)):    
    rows.append(columns_rows[row][1])

#print(rows)


wb3 = Workbook()    #Results Workbook
ws3 = wb3.active

ws3.title = 'SOPs'

for value2 in rows: #Add results to new Workbook
    for row in ws2.iter_rows(min_row = value2, max_row = value2, values_only = True):
        ws3.append(row)

wb3.save(filename = 'SOPs_results.xlsx')    #Save new Workbook

print('Created SOPs_results.xlsx')

xl = pd.ExcelFile('SOPs_results.xlsx')  #Create txt with results
file = pd.read_excel(xl,sheet_name = 'SOPs')
file.to_csv('SOPs.txt',index=False,sep='\t')
print('Created SOPs.txt')



